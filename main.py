#!/usr/bin/env python3
"""
cg-forms — Automatización de documentos PDF y Word desde CSV o Excel.

Uso:
    python main.py                    # Procesar filas pendientes
    python main.py --dry-run          # Simular sin generar archivos ni marcar la fuente
    python main.py --only-index 5     # Procesar solo la fila con índice 5
    python main.py --list-pdf-fields  # Listar campos AcroForm de cada PDF
    python main.py --validate-word-vars  # Validar variables en plantillas Word
"""

import argparse
import csv
import logging
import re
import sys
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd
from docxtpl import DocxTemplate
from pypdf import PdfReader, PdfWriter

import config

# ──────────────────────────────────────────────
# LOGGING
# ──────────────────────────────────────────────


def setup_logging() -> logging.Logger:
    """Configura logging a consola y archivo."""
    log_path = Path(config.LOG_FILE)
    log_path.parent.mkdir(parents=True, exist_ok=True)

    logger = logging.getLogger("cg-forms")
    logger.setLevel(logging.DEBUG)

    # Formato
    fmt = logging.Formatter(
        "%(asctime)s | %(levelname)-7s | %(message)s", datefmt="%Y-%m-%d %H:%M:%S"
    )

    # Consola
    ch = logging.StreamHandler(sys.stdout)
    ch.setLevel(logging.INFO)
    ch.setFormatter(fmt)
    logger.addHandler(ch)

    # Archivo
    fh = logging.FileHandler(log_path, encoding="utf-8")
    fh.setLevel(logging.DEBUG)
    fh.setFormatter(fmt)
    logger.addHandler(fh)

    return logger


log = setup_logging()

# ──────────────────────────────────────────────
# UTILIDADES
# ──────────────────────────────────────────────


def to_snake_case(name: str) -> str:
    """Convierte un nombre de columna a snake_case."""
    s = str(name).strip()
    # Reemplazar espacios, guiones y caracteres especiales por _
    s = re.sub(r"[^\w\s]", "", s)
    s = re.sub(r"[\s\-]+", "_", s)
    # Insertar _ antes de mayúsculas (camelCase → camel_case)
    s = re.sub(r"([a-z])([A-Z])", r"\1_\2", s)
    return s.lower().strip("_")


def normalize_row(row: pd.Series) -> dict:
    """Convierte una fila de pandas a dict limpio: NaN → '', keys → snake_case."""
    result = {}
    for col, val in row.items():
        if str(col).startswith("__"):
            continue
        key = to_snake_case(str(col))
        if pd.isna(val) or val is None or (isinstance(val, float) and np.isnan(val)):
            result[key] = ""
        else:
            result[key] = str(val).strip()
    return result


def enrich_context(base_ctx: dict) -> dict:
    """Aplica DERIVED_FIELDS sobre el contexto base."""
    ctx = dict(base_ctx)
    for field_name, func in config.DERIVED_FIELDS.items():
        try:
            ctx[field_name] = func(ctx)
        except Exception as e:
            log.warning(f"Error calculando campo derivado '{field_name}': {e}")
            ctx[field_name] = ""
    return ctx


def apply_mapping(ctx: dict, mapping: dict) -> dict:
    """
    Genera un dict mapeado: { clave_documento: valor_del_contexto }.
    Si mapping está vacío, devuelve el contexto completo.
    """
    if not mapping:
        return dict(ctx)

    mapped = {}
    for doc_field, ctx_field in mapping.items():
        if ctx_field in ctx:
            mapped[doc_field] = ctx[ctx_field]
        else:
            log.warning(
                f"Campo de contexto '{ctx_field}' no encontrado para '{doc_field}'. "
                f"Se usará cadena vacía."
            )
            mapped[doc_field] = ""
    return mapped


# ──────────────────────────────────────────────
# LECTURA DE EXCEL
# ──────────────────────────────────────────────


def read_excel() -> pd.DataFrame:
    """Lee el archivo fuente (Excel o CSV) y devuelve el DataFrame completo."""
    source_path = Path(config.EXCEL_SETTINGS["file"])
    if not source_path.exists():
        log.error(f"Archivo fuente no encontrado: {source_path.resolve()}")
        sys.exit(1)

    suffix = source_path.suffix.lower()
    if suffix == ".csv":
        df = pd.read_csv(source_path, dtype=str, keep_default_na=False, encoding="utf-8-sig")
        log.info(f"CSV cargado: {len(df)} filas, {len(df.columns)} columnas")
    else:
        df = pd.read_excel(
            source_path,
            sheet_name=config.EXCEL_SETTINGS["sheet_name"],
            dtype=str,
        )
        log.info(f"Excel cargado: {len(df)} filas, {len(df.columns)} columnas")

    # conservar el número de fila original del archivo para poder reescribirlo luego
    df["__source_row_number__"] = range(1, len(df) + 1)

    # eliminar filas completamente vacías y reindexar para procesamiento interno
    df = (
        df.replace(r"^\s*$", pd.NA, regex=True)
        .dropna(how="all")
        .fillna("")
        .reset_index(drop=True)
    )
    return df


def get_pending_indices(df: pd.DataFrame) -> list[int]:
    """Devuelve los índices de filas pendientes de procesar."""
    status_col = config.EXCEL_SETTINGS["status_column"]
    pending_values = config.EXCEL_SETTINGS["pending_values"]

    normalized_columns = {str(col).strip().lower(): col for col in df.columns}
    real_status_col = normalized_columns.get(status_col.strip().lower())

    if real_status_col is None:
        log.warning(
            f"Columna '{status_col}' no encontrada en la fuente de datos. "
            f"Se procesarán TODAS las filas."
        )
        return list(df.index)

    indices = []
    pending_normalized = {
        "" if v is None else str(v).strip().lower() for v in pending_values
    }

    for idx, val in df[real_status_col].items():
        # Normalizar: NaN, None, o string que coincida con pending_values
        if pd.isna(val) or val is None:
            indices.append(idx)
        elif str(val).strip().lower() in pending_normalized:
            indices.append(idx)
    return indices


def mark_as_processed(row_index: int, source_row_number: int | None = None) -> None:
    """Marca una fila como procesada en el archivo fuente original."""
    source_path = Path(config.EXCEL_SETTINGS["file"])
    status_col = config.EXCEL_SETTINGS["status_column"]
    processed_val = config.EXCEL_SETTINGS["processed_value"]

    if source_path.suffix.lower() == ".csv":
        with open(source_path, newline="", encoding="utf-8-sig") as fh:
            rows = list(csv.reader(fh))

        if not rows:
            log.error("El CSV está vacío; no se pudo marcar como procesado.")
            return

        headers = rows[0]
        normalized_headers = {str(h).strip().lower(): i for i, h in enumerate(headers)}
        col_idx = normalized_headers.get(status_col.strip().lower())

        if col_idx is None:
            log.error(f"No se encontró la columna '{status_col}' en el CSV para marcar.")
            return

        csv_row = source_row_number if source_row_number is not None else row_index + 1
        if csv_row >= len(rows):
            log.error(f"Índice de fila {row_index} fuera de rango al actualizar el CSV.")
            return

        target_row = rows[csv_row]
        if len(target_row) <= col_idx:
            target_row.extend([""] * (col_idx + 1 - len(target_row)))
        target_row[col_idx] = processed_val

        with open(source_path, "w", newline="", encoding="utf-8-sig") as fh:
            writer = csv.writer(fh)
            writer.writerows(rows)

        log.debug(f"Fila {row_index} marcada como '{processed_val}' en CSV.")
        return

    wb = openpyxl.load_workbook(source_path)
    ws = wb[config.EXCEL_SETTINGS["sheet_name"]]

    # Buscar índice de la columna de estado
    col_idx = None
    for col_num, cell in enumerate(ws[1], start=1):
        if cell.value == status_col:
            col_idx = col_num
            break

    if col_idx is None:
        log.error(f"No se encontró la columna '{status_col}' en el Excel para marcar.")
        return

    # row_index es 0-based de pandas, Excel es 1-based + 1 por header
    excel_row = (source_row_number + 1) if source_row_number is not None else row_index + 2
    ws.cell(row=excel_row, column=col_idx, value=processed_val)
    wb.save(source_path)
    log.debug(f"Fila {row_index} marcada como '{processed_val}' en Excel.")


# ──────────────────────────────────────────────
# LLENADO DE PDF
# ──────────────────────────────────────────────


def list_pdf_fields(pdf_path: str | Path) -> list[str]:
    """Lista todos los campos de formulario AcroForm de un PDF."""
    reader = PdfReader(str(pdf_path))
    fields = []
    if reader.get_fields():
        for field_name, field_obj in reader.get_fields().items():
            field_type = field_obj.get("/FT", "desconocido")
            field_value = field_obj.get("/V", "")
            fields.append(
                {
                    "name": field_name,
                    "type": str(field_type),
                    "current_value": str(field_value) if field_value else "",
                }
            )
    return fields


def fill_pdf(template_path: str | Path, output_path: str | Path, data: dict) -> None:
    """Llena un PDF con campos AcroForm usando pypdf."""
    template_path = Path(template_path)
    output_path = Path(output_path)

    if not template_path.exists():
        raise FileNotFoundError(f"Plantilla PDF no encontrada: {template_path}")

    reader = PdfReader(str(template_path))
    writer = PdfWriter()

    # Copiar todas las páginas
    for page in reader.pages:
        writer.add_page(page)

    # Obtener campos existentes para validación
    existing_fields = set()
    if reader.get_fields():
        existing_fields = set(reader.get_fields().keys())

    # Verificar campos del mapping que no existen en el PDF
    for field_name in data:
        if field_name not in existing_fields:
            log.warning(
                f"Campo '{field_name}' del mapping no existe en {template_path.name}. "
                f"Será ignorado."
            )

    # Llenar campos (pypdf ignora silenciosamente campos que no existen)
    writer.update_page_form_field_values(writer.pages[0], data)

    # Si el PDF tiene más de una página con campos, intentar llenar cada una
    if len(writer.pages) > 1:
        for page_idx in range(1, len(writer.pages)):
            try:
                writer.update_page_form_field_values(writer.pages[page_idx], data)
            except Exception:
                pass  # No todas las páginas tienen campos

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "wb") as f:
        writer.write(f)

    log.debug(f"PDF generado: {output_path}")


# ──────────────────────────────────────────────
# LLENADO DE WORD
# ──────────────────────────────────────────────


def get_word_template_vars(template_path: str | Path) -> set[str]:
    """Extrae las variables {{ var }} de una plantilla docxtpl."""
    template_path = Path(template_path)
    if not template_path.exists():
        raise FileNotFoundError(
            f"Plantilla Word no encontrada: {template_path}"
        )

    doc = DocxTemplate(str(template_path))
    return doc.get_undeclared_template_variables()


def fill_word(
    template_path: str | Path, output_path: str | Path, data: dict
) -> None:
    """Renderiza una plantilla Word con docxtpl."""
    template_path = Path(template_path)
    output_path = Path(output_path)

    if not template_path.exists():
        raise FileNotFoundError(
            f"Plantilla Word no encontrada: {template_path}"
        )

    doc = DocxTemplate(str(template_path))

    # Detectar variables usadas en la plantilla
    template_vars = doc.get_undeclared_template_variables()

    # Verificar variables faltantes en el contexto
    for var in template_vars:
        if var not in data:
            log.warning(
                f"Variable '{{{{{var}}}}}' en {template_path.name} "
                f"no tiene valor en el contexto. Se usará cadena vacía."
            )
            data[var] = ""

    output_path.parent.mkdir(parents=True, exist_ok=True)
    doc.render(data)
    doc.save(str(output_path))

    log.debug(f"Word generado: {output_path}")


# ──────────────────────────────────────────────
# PROCESAMIENTO DE UNA FILA
# ──────────────────────────────────────────────


def process_row(
    row_index: int, ctx: dict, output_dir: Path, dry_run: bool = False
) -> bool:
    """
    Procesa una fila: genera los 3 PDFs y los 2 Word.
    Retorna True si todos se generaron correctamente.
    """
    folder_name = config.output_folder_name(ctx, row_index)
    row_output = output_dir / folder_name
    row_output.mkdir(parents=True, exist_ok=True)

    total_docs = len(config.PDF_TEMPLATES) + len(config.WORD_TEMPLATES)
    generated = 0

    # ── PDFs ──
    for pdf_cfg in config.PDF_TEMPLATES:
        try:
            mapped_data = apply_mapping(ctx, pdf_cfg["mapping"])
            output_path = row_output / pdf_cfg["output_name"]

            if dry_run:
                log.info(
                    f"  [DRY-RUN] Generaría PDF: {pdf_cfg['output_name']} "
                    f"con {len(mapped_data)} campos"
                )
                generated += 1
                continue

            fill_pdf(pdf_cfg["template"], output_path, mapped_data)
            log.info(f"  ✓ PDF: {pdf_cfg['output_name']}")
            generated += 1

        except Exception as e:
            log.error(
                f"  ✗ Error generando PDF '{pdf_cfg['output_name']}': {e}",
                exc_info=True,
            )

    # ── Word ──
    for word_cfg in config.WORD_TEMPLATES:
        try:
            mapped_data = apply_mapping(ctx, word_cfg["mapping"])
            output_path = row_output / word_cfg["output_name"]

            if dry_run:
                log.info(
                    f"  [DRY-RUN] Generaría Word: {word_cfg['output_name']} "
                    f"con {len(mapped_data)} variables"
                )
                generated += 1
                continue

            fill_word(word_cfg["template"], output_path, mapped_data)
            log.info(f"  ✓ Word: {word_cfg['output_name']}")
            generated += 1

        except Exception as e:
            log.error(
                f"  ✗ Error generando Word '{word_cfg['output_name']}': {e}",
                exc_info=True,
            )

    return generated == total_docs


# ──────────────────────────────────────────────
# UTILIDADES CLI
# ──────────────────────────────────────────────


def cmd_list_pdf_fields() -> None:
    """Lista los campos AcroForm de todas las plantillas PDF configuradas."""
    log.info("=" * 60)
    log.info("CAMPOS DE FORMULARIO PDF")
    log.info("=" * 60)

    for pdf_cfg in config.PDF_TEMPLATES:
        template_path = Path(pdf_cfg["template"])
        log.info(f"\n📄 {template_path}")

        if not template_path.exists():
            log.error(f"   Archivo no encontrado: {template_path.resolve()}")
            continue

        fields = list_pdf_fields(template_path)
        if not fields:
            log.warning("   No se encontraron campos AcroForm en este PDF.")
            log.info(
                "   (Asegúrate de que el PDF tenga campos de formulario rellenables)"
            )
            continue

        log.info(f"   {len(fields)} campo(s) encontrado(s):")
        for f in fields:
            log.info(
                f'   - "{f["name"]}" (tipo: {f["type"]}, '
                f'valor actual: "{f["current_value"]}")'
            )

    log.info("\n" + "=" * 60)
    log.info(
        "Copia los nombres de campo exactos en el mapping de config.py"
    )
    log.info("=" * 60)


def cmd_validate_word_vars() -> None:
    """Valida las variables de cada plantilla Word contra el mapping."""
    log.info("=" * 60)
    log.info("VALIDACIÓN DE VARIABLES WORD")
    log.info("=" * 60)

    for word_cfg in config.WORD_TEMPLATES:
        template_path = Path(word_cfg["template"])
        log.info(f"\n📝 {template_path}")

        if not template_path.exists():
            log.error(f"   Archivo no encontrado: {template_path.resolve()}")
            continue

        try:
            template_vars = get_word_template_vars(template_path)
        except Exception as e:
            log.error(f"   Error leyendo plantilla: {e}")
            continue

        if not template_vars:
            log.warning(
                "   No se encontraron variables {{...}} en esta plantilla."
            )
            continue

        log.info(f"   {len(template_vars)} variable(s) en la plantilla:")
        mapping = word_cfg.get("mapping", {})

        for var in sorted(template_vars):
            if mapping:
                if var in mapping:
                    log.info(
                        f'   ✓ "{var}" → mapeada a "{mapping[var]}"'
                    )
                else:
                    log.warning(
                        f'   ⚠ "{var}" → NO está en el mapping '
                        f"(se buscará directamente en el contexto)"
                    )
            else:
                log.info(
                    f'   - "{var}" (sin mapping, se usa contexto directo)'
                )

        # Variables en mapping que no están en la plantilla
        if mapping:
            extra = set(mapping.keys()) - template_vars
            if extra:
                log.warning(
                    f"   Variables en mapping que NO están en la plantilla: {extra}"
                )

    log.info("\n" + "=" * 60)


# ──────────────────────────────────────────────
# ORQUESTADOR PRINCIPAL
# ──────────────────────────────────────────────


def run(
    dry_run: bool = False,
    only_index: int | None = None,
) -> None:
    """Flujo principal de procesamiento."""
    log.info("=" * 60)
    log.info("cg-forms — Inicio de procesamiento")
    log.info(f"Modo: {'DRY-RUN (simulación)' if dry_run else 'PRODUCCIÓN'}")
    log.info("=" * 60)

    # 1. Leer fuente de datos
    df = read_excel()

    # 2. Obtener filas pendientes
    pending = get_pending_indices(df)

    if only_index is not None:
        if only_index not in pending:
            log.warning(
                f"Fila {only_index} no está pendiente o no existe. "
                f"Procesando de todas formas."
            )
            if only_index >= len(df):
                log.error(f"Índice {only_index} fuera de rango (total: {len(df)} filas)")
                return
        pending = [only_index]

    log.info(f"Filas pendientes: {len(pending)} de {len(df)} totales")

    if not pending:
        log.info("No hay filas pendientes. Nada que procesar.")
        return

    output_dir = Path(config.OUTPUT_DIR)
    output_dir.mkdir(parents=True, exist_ok=True)

    # Contadores
    exitosos = 0
    fallidos = 0

    # 3. Procesar cada fila
    for idx in pending:
        row = df.iloc[idx]
        source_row_number = row.get("__source_row_number__")
        base_ctx = normalize_row(row)
        ctx = enrich_context(base_ctx)

        log.info(f"\n{'─' * 40}")
        log.info(
            f"Procesando fila {idx} | "
            f"{config.output_folder_name(ctx, idx)}"
        )
        log.info(f"{'─' * 40}")

        success = process_row(idx, ctx, output_dir, dry_run=dry_run)

        if success:
            if not dry_run:
                try:
                    if source_row_number not in (None, ""):
                        mark_as_processed(idx, int(source_row_number))
                    else:
                        mark_as_processed(idx)
                    log.info(f"  ✓ Fila {idx} marcada como procesada")
                except Exception as e:
                    log.error(
                        f"  ✗ Error marcando fila {idx} como procesada: {e}",
                        exc_info=True,
                    )
                    fallidos += 1
                    continue
            else:
                log.info(f"  [DRY-RUN] Se marcaría fila {idx} como procesada")
            exitosos += 1
        else:
            log.error(
                f"  ✗ Fila {idx} tuvo errores. NO se marca como procesada."
            )
            fallidos += 1

    # 4. Resumen final
    log.info(f"\n{'=' * 60}")
    log.info("RESUMEN FINAL")
    log.info(f"{'=' * 60}")
    log.info(f"Total filas en fuente:     {len(df)}")
    log.info(f"Filas pendientes al inicio: {len(pending)}")
    log.info(f"Exitosas:                  {exitosos}")
    log.info(f"Fallidas:                  {fallidos}")
    log.info(
        f"Pendientes restantes:      {len(pending) - exitosos - fallidos + fallidos}"
    )
    log.info(f"Modo:                      {'DRY-RUN' if dry_run else 'PRODUCCIÓN'}")
    log.info(f"Log guardado en:           {config.LOG_FILE}")
    log.info("=" * 60)


# ──────────────────────────────────────────────
# CLI
# ──────────────────────────────────────────────


def parse_args() -> argparse.Namespace:
    """Define y parsea argumentos de línea de comandos."""
    parser = argparse.ArgumentParser(
        description="cg-forms — Automatización de documentos PDF y Word desde CSV o Excel",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Ejemplos:
  python main.py                        Procesar todas las filas pendientes
  python main.py --dry-run              Simular sin generar archivos
  python main.py --only-index 3         Procesar solo la fila con índice 3
  python main.py --list-pdf-fields      Listar campos de los PDFs
  python main.py --validate-word-vars   Validar variables en plantillas Word
        """,
    )

    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Simular el procesamiento sin generar archivos ni modificar la fuente de datos",
    )
    parser.add_argument(
        "--only-index",
        type=int,
        default=None,
        help="Procesar únicamente la fila con este índice (0-based)",
    )
    parser.add_argument(
        "--list-pdf-fields",
        action="store_true",
        help="Listar los campos AcroForm de cada plantilla PDF y salir",
    )
    parser.add_argument(
        "--validate-word-vars",
        action="store_true",
        help="Validar variables de cada plantilla Word contra el mapping y salir",
    )

    return parser.parse_args()


def main() -> None:
    """Punto de entrada principal."""
    args = parse_args()

    if args.list_pdf_fields:
        cmd_list_pdf_fields()
        return

    if args.validate_word_vars:
        cmd_validate_word_vars()
        return

    run(dry_run=args.dry_run, only_index=args.only_index)


if __name__ == "__main__":
    main()
